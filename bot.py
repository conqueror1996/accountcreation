import os
import random
import re
import threading
import time
import urllib.request
import urllib.parse
import json
import string
import requests
from bs4 import BeautifulSoup
from http.server import BaseHTTPRequestHandler, HTTPServer
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from dotenv import load_dotenv

# सुरक्षा और आसानी के लिए टोकन लोड करना
load_dotenv()
OTP_API_KEY = os.getenv("OTP_API_KEY", "emsy9uqwebdwvmwmqe8n47rlk3530ehv")

# ---------------- STORAGE & STATE ----------------
sessions = {}
EXCEL_FILE = "users.xlsx"
excel_lock = threading.Lock()

is_automation_running = {}  # ट्रैक रखने के लिए कि किस यूजर का लूप चल रहा है

# ---------------- DASHBOARD LOGGING ----------------
import builtins
dashboard_logs = []

def add_log(message):
    timestamp = time.strftime("[%H:%M:%S]")
    log_line = f"{timestamp} {message}"
    builtins.print(log_line)
    dashboard_logs.append(log_line)
    if len(dashboard_logs) > 50:
        dashboard_logs.pop(0)

# Redefine standard print in this namespace
def print(*args, **kwargs):
    msg = " ".join(str(arg) for arg in args)
    add_log(msg)


# ---------------- HTTP SERVER FOR DASHBOARD ----------------
class DashboardHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        # Suppress logging in CLI to keep console output clean
        pass

    def do_GET(self):
        if self.path == "/":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            current_dir = os.path.dirname(os.path.abspath(__file__))
            dashboard_path = os.path.join(current_dir, "dashboard.html")
            try:
                if os.path.exists(dashboard_path):
                    with open(dashboard_path, "r", encoding="utf-8") as f:
                        self.wfile.write(f.read().encode("utf-8"))
                else:
                    self.wfile.write("<h1>dashboard.html not found!</h1>".encode("utf-8"))
            except Exception as e:
                self.wfile.write(f"<h1>Error loading dashboard: {e}</h1>".encode("utf-8"))
        elif self.path == "/api/status":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            
            user_id = list(sessions.keys())[0] if sessions else None
            is_running = any(is_automation_running.values())
            
            # Read dynamic API key from headers
            api_key = self.headers.get("X-OTP-API-Key")
            balance = get_otp_balance(api_key)
            
            urls = []
            service_id = "10704"
            if user_id and user_id in sessions:
                urls = sessions[user_id].get("urls", [])
                if not urls and "url" in sessions[user_id]:
                    urls = [sessions[user_id]["url"]]
                service_id = sessions[user_id].get("service_id", "10704")
                
            status_data = {
                "is_running": is_running,
                "balance": balance if balance is not None else 0.0,
                "urls": urls,
                "service_id": service_id,
                "logs": dashboard_logs
            }
            self.wfile.write(json.dumps(status_data).encode("utf-8"))
        elif self.path == "/api/data":
            if os.path.exists(EXCEL_FILE):
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f"attachment; filename={EXCEL_FILE}")
                self.end_headers()
                with open(EXCEL_FILE, "rb") as f:
                    self.wfile.write(f.read())
            else:
                self.send_response(404)
                self.end_headers()
        elif self.path == "/api/users":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            users = get_all_registered_users()
            self.wfile.write(json.dumps(users).encode("utf-8"))
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        content_length = int(self.headers.get('Content-Length', 0))
        if content_length > 0:
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode("utf-8"))
        else:
            data = {}
        
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        
        user_id = list(sessions.keys())[0] if sessions else 999999999
        
        if self.path == "/api/start":
            urls = data.get("urls", [])
            service_id = data.get("service_id", "10704")
            otp_api_key = data.get("otp_api_key", "")
            
            if user_id not in sessions:
                sessions[user_id] = {}
            sessions[user_id]["urls"] = urls
            sessions[user_id]["service_id"] = service_id
            sessions[user_id]["otp_api_key"] = otp_api_key
            
            if not is_automation_running.get(user_id, False):
                is_automation_running[user_id] = True
                threads_count = int(data.get("threads", 1))
                for i in range(threads_count):
                    threading.Thread(target=start_automation_loop, args=(user_id, i+1), daemon=True).start()
                add_log(f"Dashboard: Automation started with {threads_count} threads.")
                
            self.wfile.write(json.dumps({"status": "success", "message": "Automation started"}).encode("utf-8"))
            
        elif self.path == "/api/stop":
            if user_id in is_automation_running:
                is_automation_running[user_id] = False
                add_log("Dashboard: Automation stop requested.")
            self.wfile.write(json.dumps({"status": "success", "message": "Automation stop requested"}).encode("utf-8"))
            
        elif self.path == "/api/user/balance":
            domain = data.get("domain")
            username = data.get("username")
            password = data.get("password")
            balance = fetch_live_balance(domain, username, password)
            self.wfile.write(json.dumps({"status": "success", "balance": balance}).encode("utf-8"))
            
        elif self.path == "/api/user/deposit/get-upi":
            domain = data.get("domain")
            username = data.get("username")
            password = data.get("password")
            amount = data.get("amount", 500)
            res = fetch_deposit_upi(domain, username, password, amount)
            self.wfile.write(json.dumps(res).encode("utf-8"))
            
        elif self.path == "/api/user/deposit/submit-utr":
            domain = data.get("domain")
            username = data.get("username")
            password = data.get("password")
            utr = data.get("utr")
            amount = data.get("amount", 500)
            res = submit_deposit_utr(domain, username, password, utr, amount)
            self.wfile.write(json.dumps(res).encode("utf-8"))

        elif self.path == "/api/user/delete":
            domain = data.get("domain", "")
            username = data.get("username", "")
            success = delete_user_from_excel(domain, username)
            if success:
                self.wfile.write(json.dumps({"status": "success", "message": f"Deleted {username} from {domain}"}).encode("utf-8"))
            else:
                self.wfile.write(json.dumps({"status": "error", "message": "User not found"}).encode("utf-8"))

        elif self.path == "/api/user/logout-all":
            active_sessions = []
            with sessions_lock:
                active_sessions = list(authenticated_sessions.items())
                
            # Trigger explicit server-side logout request on each casino host
            for key, (session, _) in active_sessions:
                domain, username = key
                if not domain.startswith("http"):
                    url = "https://" + domain
                else:
                    url = domain
                try:
                    csrf = getattr(session, "csrf_token", None)
                    if csrf:
                        session.post(f"{url.rstrip('/')}/logout", data={"_token": csrf}, timeout=5)
                    else:
                        session.get(f"{url.rstrip('/')}/logout", timeout=5)
                except Exception:
                    pass
                    
            with sessions_lock:
                authenticated_sessions.clear()
            add_log("🔓 All cached active browser sessions cleared successfully")
            self.wfile.write(json.dumps({"status": "success", "message": "All cached sessions cleared successfully"}).encode("utf-8"))


# ---------------- OTP DOCTOR SERVICE MAPPINGS ----------------
DEFAULT_SERVICES = {
    "cricmatch": "6272",   # Cricmatch
    "starexch": "10940",   # Starexch
    "khelstake": "12827",  # Any Other
    "playcric": "6272",    # Cricmatch or Any Other (default 6272)
    "playkaro": "10704"    # Playkaro
}

# ---------------- OTP DOCTOR API HELPERS ----------------
cached_balances = {}  # api_key -> (balance, last_check)
last_balance_error_time = 0.0

def get_otp_balance(api_key=None):
    global last_balance_error_time
    if not api_key:
        api_key = OTP_API_KEY
    if not api_key:
        return 0.0

    now = time.time()
    if api_key in cached_balances:
        val, last_check = cached_balances[api_key]
        if now - last_check < 15:
            return val

    # Limit check error frequency on network failures
    if now - last_balance_error_time < 30 and last_balance_error_time > 0:
        if api_key in cached_balances:
            return cached_balances[api_key][0]
        return 0.0

    url = f"https://www.otpdoctor.in/stubs/handler_api.php?action=getBalance&api_key={api_key}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            res = response.read().decode('utf-8').strip()
            if res.startswith("ACCESS_BALANCE:"):
                val = float(res.split(":")[1])
                cached_balances[api_key] = (val, now)
                last_balance_error_time = 0.0
                return val
            return 0.0
    except Exception as e:
        last_balance_error_time = now
        if api_key in cached_balances:
            return cached_balances[api_key][0]
        return 0.0

def request_otp_number(service_id, api_key=None):
    if not api_key:
        api_key = OTP_API_KEY
    if not api_key:
        return {"status": "error", "message": "API key not configured"}
        
    url = f"https://www.otpdoctor.in/stubs/handler_api.php?action=getNumber&api_key={api_key}&service={service_id}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as response:
            res = response.read().decode('utf-8').strip()
            if res.startswith("ACCESS_NUMBER:"):
                parts = res.split(":")
                return {
                    "status": "success",
                    "activation_id": parts[1],
                    "phone": parts[2]
                }
            return {
                "status": "error",
                "message": res
            }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def get_otp_status(activation_id, api_key=None):
    if not api_key:
        api_key = OTP_API_KEY
    if not api_key:
        return "ERROR"
        
    url = f"https://www.otpdoctor.in/stubs/handler_api.php?action=getStatus&api_key={api_key}&id={activation_id}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            res = response.read().decode('utf-8').strip()
            return res
    except Exception as e:
        print(f"Error getting OTP status for {activation_id}: {e}")
        return "ERROR"

def set_otp_status(activation_id, status, api_key=None):
    if not api_key:
        api_key = OTP_API_KEY
    if not api_key:
        return "ERROR"
        
    url = f"https://www.otpdoctor.in/stubs/handler_api.php?action=setStatus&api_key={api_key}&id={activation_id}&status={status}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            res = response.read().decode('utf-8').strip()
            return res
    except Exception as e:
        print(f"Error setting OTP status for {activation_id} to {status}: {e}")
        return "ERROR"

def extract_otp_code(sms_text, used_otps=None):
    if used_otps is None:
        used_otps = set()
        
    # Find all 6 digit numbers
    six_digits = re.findall(r"\b\d{6}\b", sms_text)
    # Find all 4 digit numbers
    four_digits = re.findall(r"\b\d{4}\b", sms_text)
    
    # Check 6 digits first, starting from the newest (reverse order)
    for code in reversed(six_digits):
        if code not in used_otps:
            return code
            
    # Then check 4 digits, starting from newest
    for code in reversed(four_digits):
        if code not in used_otps:
            return code
            
    return None


# ---------------- EXCEL INIT ----------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.save(EXCEL_FILE)


def save_user(domain, username, password, phone):
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    headers = ["Timestamp", "Domain", "Username", "Password", "Phone"]
    
    with excel_lock:
        try:
            wb = load_workbook(EXCEL_FILE)
        except Exception:
            wb = Workbook()

        # Clean sheet name (Excel sheet name limit is 31 chars)
        sheet_name = domain.strip().lower()
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        # Get or create the sheet for this domain
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)

        ws.append([timestamp, domain, username, password, phone])

        # Remove default empty "Sheet" or "Registered Accounts" if they exist
        for default_name in ["Sheet", "Registered Accounts"]:
            if default_name in wb.sheetnames and len(wb.sheetnames) > 1:
                default_sheet = wb[default_name]
                if default_sheet.max_row <= 1:
                    wb.remove(default_sheet)

        wb.save(EXCEL_FILE)


init_excel()


# ---------------- USER INFO & PAYMENT HELPERS ----------------
from bs4 import BeautifulSoup

def get_all_registered_users():
    users_by_domain = {}
    if not os.path.exists(EXCEL_FILE):
        return users_by_domain
        
    try:
        wb = load_workbook(EXCEL_FILE)
        for sheet_name in wb.sheetnames:
            if sheet_name == "Sheet":
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))[1:]
            for row in rows:
                if not row or len(row) < 5 or not row[2]:
                    continue
                domain_val = str(row[1]).strip().lower()
                if not domain_val:
                    continue
                user_obj = {
                    "timestamp": str(row[0]),
                    "domain": str(row[1]),
                    "username": str(row[2]),
                    "password": str(row[3]),
                    "phone": str(row[4]),
                    "balance": "Click Refresh"
                }
                if domain_val not in users_by_domain:
                    users_by_domain[domain_val] = []
                users_by_domain[domain_val].append(user_obj)
    except Exception as e:
        print(f"Error reading users from Excel: {e}")
    return users_by_domain


def delete_user_from_excel(domain, username):
    """Delete a user row from the Excel file by domain and username."""
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        wb = load_workbook(EXCEL_FILE)
        found = False
        for sheet_name in wb.sheetnames:
            if sheet_name == "Sheet":
                continue
            ws = wb[sheet_name]
            for row_idx in range(ws.max_row, 1, -1):  # reverse to safely delete
                row_domain = str(ws.cell(row=row_idx, column=2).value or "").strip().lower()
                row_user = str(ws.cell(row=row_idx, column=3).value or "").strip()
                if row_domain == domain.strip().lower() and row_user == username.strip():
                    ws.delete_rows(row_idx)
                    found = True
                    break
            if found:
                break
        if found:
            wb.save(EXCEL_FILE)
            add_log(f"Deleted user '{username}' from domain '{domain}'")
        return found
    except Exception as e:
        print(f"Error deleting user from Excel: {e}")
        return False


authenticated_sessions = {}
sessions_lock = threading.Lock()

def get_authenticated_session(domain, username, password):
    if not domain.startswith("http"):
        url = "https://" + domain
    else:
        url = domain
        
    key = (domain, username)
    
    with sessions_lock:
        if key in authenticated_sessions:
            session, last_time = authenticated_sessions[key]
            # If session was created/accessed recently (within 15 minutes), reuse it
            if time.time() - last_time < 900:
                authenticated_sessions[key] = (session, time.time())
                return session, None
                
    # Create a fresh session
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Origin": url,
        "Referer": url + "/"
    })
    
    try:
        r_home = session.get(url, timeout=15)
        token = None
        match = re.search(r'name="csrf-token"\s+content="([^"]+)"', r_home.text)
        if not match:
            match = re.search(r'value="([^"]+)"[^>]*name="_token"', r_home.text)
        if match:
            token = match.group(1)
        if not token:
            return None, "CSRF not found"
            
        payload = {
            "username": username,
            "password": password,
            "remember_me": 1,
            "_token": token
        }
        r_login = session.post(f"{url.rstrip('/')}/login", data=payload, timeout=15)
        if r_login.status_code != 200 or "invalid" in r_login.text.lower():
            return None, "Login failed on game server"
            
        login_data = r_login.json()
        uid_match = re.search(r'uid=(\d+)', login_data.get("url", ""))
        uid = uid_match.group(1) if uid_match else "7260553"
        
        # Store attributes on session object for reuse
        session.csrf_token = token
        session.uid = uid
        
        # Access home page as logged-in session context
        target_url = f"{url.rstrip('/')}/?uid={uid}"
        session.get(target_url, timeout=15)
        
        with sessions_lock:
            authenticated_sessions[key] = (session, time.time())
            
        return session, None
    except Exception as e:
        return None, str(e)


def fetch_live_balance(domain, username, password):
    if not domain.startswith("http"):
        url = "https://" + domain
    else:
        url = domain
        
    try:
        # Try retrieving active session from cache
        session, err = get_authenticated_session(domain, username, password)
        if err:
            return f"Error ({err[:12]})"
            
        # Post request to getBalance endpoint
        session.headers.update({
            "X-CSRF-Token": session.csrf_token,
            "X-Requested-With": "XMLHttpRequest",
            "Referer": f"{url.rstrip('/')}/?uid={session.uid}"
        })
        r_bal = session.post(f"{url.rstrip('/')}/api2/v2/getBalance", data={"_token": session.csrf_token}, timeout=10)
        
        # If session expired on the server, invalidate cache and retry once
        if r_bal.status_code != 200:
            with sessions_lock:
                authenticated_sessions.pop((domain, username), None)
            session, err = get_authenticated_session(domain, username, password)
            if err:
                return f"Error ({err[:12]})"
            session.headers.update({
                "X-CSRF-Token": session.csrf_token,
                "X-Requested-With": "XMLHttpRequest",
                "Referer": f"{url.rstrip('/')}/?uid={session.uid}"
            })
            r_bal = session.post(f"{url.rstrip('/')}/api2/v2/getBalance", data={"_token": session.csrf_token}, timeout=10)
            
        if r_bal.status_code == 200:
            bal_data = r_bal.json()
            if "balance" in bal_data and "totalBalance" in bal_data["balance"]:
                val = bal_data["balance"]["totalBalance"]
                return f"₹{float(val):,.2f}"
            elif "balance" in bal_data and "balance" in bal_data["balance"]:
                val = bal_data["balance"]["balance"]
                return f"₹{float(val):,.2f}"
                
        return "₹0.00"
    except Exception as e:
        return f"Error ({str(e)[:12]})"


def fetch_deposit_upi(domain, username, password, amount=500):
    if not domain.startswith("http"):
        url = "https://" + domain
    else:
        url = domain
        
    try:
        # Try retrieving active session from cache
        session, err = get_authenticated_session(domain, username, password)
        if err:
            return {"status": "error", "message": f"Session error: {err}"}
            
        # POST to /pay/v3/store
        store_payload = {
            "hashed": "",
            "amount": str(amount),
            "coupon": "",
            "userid": session.uid,
            "link": "https://stipepay.com/pg/stipeyWNO/checkout",
            "_token": session.csrf_token
        }
        
        # Update headers for AJAX
        session.headers.update({
            "X-CSRF-Token": session.csrf_token,
            "X-Requested-With": "XMLHttpRequest",
            "Referer": f"{url.rstrip('/')}/?uid={session.uid}"
        })
        r_store = session.post(f"{url.rstrip('/')}/pay/v3/store", data=store_payload, timeout=15)
        
        # Self-healing: if session expired, clear cache, re-authenticate and retry
        if r_store.status_code != 200 or not r_store.json().get("key"):
            with sessions_lock:
                authenticated_sessions.pop((domain, username), None)
            session, err = get_authenticated_session(domain, username, password)
            if err:
                return {"status": "error", "message": f"Session error: {err}"}
            store_payload["_token"] = session.csrf_token
            session.headers.update({
                "X-CSRF-Token": session.csrf_token,
                "X-Requested-With": "XMLHttpRequest",
                "Referer": f"{url.rstrip('/')}/?uid={session.uid}"
            })
            r_store = session.post(f"{url.rstrip('/')}/pay/v3/store", data=store_payload, timeout=15)
            
        store_json = r_store.json()
        hashed_key = store_json.get("key")
        if not hashed_key:
            return {"status": "error", "message": "Failed to get store hash"}
            
        # 4. POST to stipepay checkout
        session.headers.pop("X-Requested-With", None)
        session.headers.pop("X-CSRF-Token", None)
        submit_payload = {
            "hashed": hashed_key,
            "amount": str(amount),
            "coupon": "",
            "userid": session.uid,
            "link": "https://stipepay.com/pg/stipeyWNO/checkout"
        }
        r_submit = session.post("https://stipepay.com/pg/stipeyWNO/checkout", data=submit_payload, timeout=15)
        
        # 5. Parse first redirect form
        soup1 = BeautifulSoup(r_submit.text, 'html.parser')
        form1 = soup1.select_one("#payForm")
        if not form1:
            return {"status": "error", "message": "Stipepay redirection failed"}
            
        # Parse UPI ID for display on UI card
        upi_id = None
        details = soup1.select(".accountInfos")
        for d in details:
            try:
                label = d.select_one(".accountInfo").text.strip()
                val = d.select_one(".accountinfoDetails").text.strip()
                if "upi" in label.lower():
                    upi_id = val
                    break
            except:
                pass
        if not upi_id:
            upi_id = "payonline@upi"
            
        action_url1 = form1.get("action")
        form_inputs1 = {}
        for ip in form1.find_all("input"):
            form_inputs1[ip.get("name")] = ip.get("value", "")
            
        # 6. POST to first redirect URL
        r_redirect1 = session.post(action_url1, data=form_inputs1, timeout=15)
        
        # 7. Parse second redirect form
        soup2 = BeautifulSoup(r_redirect1.text, 'html.parser')
        form2 = soup2.select_one("#payForm")
        if not form2:
            return {"status": "error", "message": "Truelayer redirection failed"}
            
        action_url2 = form2.get("action")
        if action_url2.startswith("/"):
            action_url2 = "https://app.truelayerpayments.com" + action_url2
            
        form_inputs2 = {}
        for ip in form2.find_all("input"):
            form_inputs2[ip.get("name")] = ip.get("value", "")
            
        return {
            "status": "success",
            "upi": upi_id,
            "action": action_url2,
            "fields": form_inputs2,
            "message": "Deposit initialized successfully"
        }
    except Exception as e:
        return {"status": "error", "message": f"Error initiating deposit: {str(e)}"}


def submit_deposit_utr(domain, username, password, utr, amount=500):
    if not domain.startswith("http"):
        url = "https://" + domain
    else:
        url = domain
        
    try:
        session, err = get_authenticated_session(domain, username, password)
        if err:
            return {"status": "error", "message": f"Session error: {err}"}
            
        # POST to depositamount UTR verification endpoint
        session.headers.update({
            "X-CSRF-Token": session.csrf_token,
            "X-Requested-With": "XMLHttpRequest",
            "Referer": f"{url.rstrip('/')}/?uid={session.uid}"
        })
        
        post_payload = {
            "depositvalue": str(amount),
            "txn_id": str(utr),
            "_token": session.csrf_token,
            "type": "UPI"
        }
        r_res = session.post(f"{url.rstrip('/')}/depositamount", data=post_payload, timeout=15)
        print(f"[Engine] UTR Submit Response: {r_res.status_code} | {r_res.text}")
        
        return {
            "status": "success",
            "message": "Deposit UTR submitted successfully! It will be verified by the game server."
        }
    except Exception as e:
        return {"status": "error", "message": f"Error submitting UTR: {str(e)}"}


# ---------------- USERNAME GENERATOR ----------------
def generate_username():
    names = ["Karan", "Amit", "Rohan", "Rahul", "Yash", "Varun", "Abhishek", "Deepak", "Vivek", "Tarun"]
    surnames = ["Sharma", "Verma", "Kumar", "Singh", "Gupta", "Yadav", "Joshi", "Mishra", "Patel", "Das"]
    random_suffix = "".join(random.choices(string.ascii_letters, k=6))
    return f"{random.choice(names)}{random.choice(surnames)}{random_suffix}"


# ---------------- DRIVER INIT ----------------
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    return webdriver.Chrome(options=options)


# ---------------- SAFE SELENIUM FUNCTIONS ----------------
def safe_click(driver, by, value):
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((by, value))
    )
    elements = driver.find_elements(by, value)
    # Try to find a visible and enabled element first
    for el in elements:
        try:
            if el.is_displayed() and el.is_enabled():
                driver.execute_script("arguments[0].click();", el)
                return
        except:
            pass
    # Fallback to the first element if none of them matched
    if elements:
        try:
            driver.execute_script("arguments[0].click();", elements[0])
        except Exception as e:
            print(f"⚠️ safe_click fallback failed: {e}")


def safe_fill(driver, by, value, text):
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((by, value))
    )
    elements = driver.find_elements(by, value)
    for el in elements:
        try:
            if el.is_displayed() and el.is_enabled():
                el.clear()
                el.send_keys(text)
                return
        except:
            pass
    # Fallback to first element and use JS if send_keys is not interactable
    if elements:
        try:
            elements[0].clear()
            elements[0].send_keys(text)
        except Exception:
            try:
                driver.execute_script("arguments[0].value = arguments[1];", elements[0], text)
                driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", elements[0])
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", elements[0])
            except Exception as e:
                print(f"⚠️ safe_fill JS fallback failed: {e}")


# ---------------- POPUP & REGISTRATION HELPERS ----------------
def dismiss_popups(driver):
    """Only close PROMOTIONAL popups, NOT the registration/signup modal."""
    selectors = [
        (By.CSS_SELECTOR, "button.mnPopupClose"),
        (By.CSS_SELECTOR, "button.pgSoftClsBtn"),
        (By.CSS_SELECTOR, "button.mnPopupBtn.mnPopupClose"),
    ]
    for by, val in selectors:
        try:
            elements = driver.find_elements(by, val)
            for el in elements:
                if el.is_displayed() and el.is_enabled():
                    driver.execute_script("arguments[0].click();", el)
                    print(f"🗑️ Dismissed promo popup: {val}")
                    time.sleep(1)
        except:
            pass

def click_register_button(driver):
    register_selectors = [
        "button.cls_register_new",
        "button.rgstrBtn",
        "button.action_btn",
        "button[type='submit']",
        ".cls_register_new",
        ".rgstrBtn"
    ]
    for sel in register_selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in elements:
                if el.is_displayed() and el.is_enabled():
                    driver.execute_script("arguments[0].click();", el)
                    print(f"Clicked register button using selector: {sel}")
                    return True
        except:
            pass
            
    # Fallback wait to raise exception if not found
    el = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "button.rgstrBtn"))
    )
    driver.execute_script("arguments[0].click();", el)
    return True

def fill_otp_and_verify(driver, extracted_otp):
    # Try to locate OTP inputs using different selectors
    otp_selectors = [
        "input.otp_v_num",
        "input.otp__digit_signup",
        "input.register_otp",
        "input.otp__digit",
        ".otp__digit"
    ]
    boxes = []
    for sel in otp_selectors:
        try:
            boxes = driver.find_elements(By.CSS_SELECTOR, sel)
            boxes = [b for b in boxes if b.is_displayed()]
            if len(boxes) >= 4:
                break
        except:
            pass
            
    if not boxes:
        # Fallback wait to raise exception if not found
        boxes = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "input.register_otp"))
        )
        
    otp_len = min(len(extracted_otp), len(boxes))
    for i in range(otp_len):
        try:
            boxes[i].click()
            boxes[i].clear()
            boxes[i].send_keys(extracted_otp[i])
        except Exception as e:
            # Fallback to JS if normal typing fails
            try:
                driver.execute_script("arguments[0].value = arguments[1];", boxes[i], extracted_otp[i])
                driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", boxes[i])
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", boxes[i])
            except:
                print(f"⚠️ Failed to fill OTP box {i}: {e}")

    # Verify बटन ढूँढना
    verify_selectors = [
        ".cls_verify_nw_signup",
        ".get_user_otp",
        ".cls_verify_otp",
        "a.get_user_otp",
        "button.get_user_otp",
        "button.cls_verify_nw_signup"
    ]
    verify_btn = None
    for sel in verify_selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in elements:
                if el.is_displayed() and el.is_enabled():
                    verify_btn = el
                    break
            if verify_btn:
                break
        except:
            pass
            
    if not verify_btn:
        verify_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.get_user_otp.cls_verify_nw_signup"))
        )
        
    # Click Verify
    try:
        driver.execute_script("arguments[0].click();", verify_btn)
    except Exception as e:
        print(f"⚠️ Failed to click verify button: {e}")
        # Try native click
        verify_btn.click()


# ---------------- AUTOMATION LOOP ----------------
def start_automation_loop(user_id, thread_id=1):
    """सुरक्षित और कंट्रोल्ड रिस्पॉन्स-वेरिफिकेशन लूप"""
    def thread_print(*args, **kwargs):
        msg = " ".join(str(arg) for arg in args)
        globals()["print"](f"[Engine-{thread_id}] {msg}")
    print = thread_print

    # Pre-populate candidate service IDs list
    candidate_services = ["6272", "10704", "10940", "12827"]
    first_run = True

    while is_automation_running.get(user_id, False):
        # Read session data
        data = sessions.get(user_id)
        if not data:
            print("❌ No session data found.")
            break
            
        urls = data.get("urls", [])
        if not urls and "url" in data:
            urls = [data["url"]]
        if not urls:
            print("❌ No URLs found in session.")
            break
            
        # Determine candidate service IDs
        selected_service_id = data.get("service_id", "auto")
        if selected_service_id == "auto":
            if first_run:
                # Priority candidates based on first target URL
                candidate_services = ["6272", "10704", "10940", "12827"]
                if urls:
                    first_url = urls[0].lower()
                    if "playkaro" in first_url:
                        candidate_services = ["10704", "12827", "6272", "10940"]
                    elif "starexch" in first_url:
                        candidate_services = ["10940", "12827", "6272", "10704"]
                    elif "cricmatch" in first_url:
                        candidate_services = ["6272", "12827", "10940", "10704"]
                    elif "khelstake" in first_url:
                        candidate_services = ["12827", "6272", "10940", "10704"]
                first_run = False
        else:
            candidate_services = [selected_service_id]

        # Read session dynamic API Key
        api_key = data.get("otp_api_key")

        # Check balance first
        balance = get_otp_balance(api_key)
        if balance is not None:
            if isinstance(balance, float) and balance < 10.0:
                print(f"❌ Insufficient OTP Doctor balance: {balance}")
                is_automation_running[user_id] = False
                break
        
        # Try candidate services one by one
        num_res = None
        used_service_id = None
        for s_id in candidate_services:
            print(f"📡 OTP Doctor से नया नंबर लिया जा रहा है (Service ID: {s_id})...")
            num_res = request_otp_number(s_id, api_key)
            if num_res.get("status") == "success":
                used_service_id = s_id
                break
            else:
                print(f"⚠️ Service ID {s_id} से नंबर नहीं मिला: {num_res.get('message', 'Unknown error')}")

        if not num_res or num_res.get("status") != "success":
            err_msg = num_res.get("message", "Unknown error") if num_res else "No service responded"
            if "NO_BALANCE" in err_msg:
                is_automation_running[user_id] = False
                break
            print(f"⚠️ किसी भी सर्विस से नंबर नहीं मिला: {err_msg}। 10 सेकंड में पुनः प्रयास किया जाएगा...")
            time.sleep(10)
            continue
            
        activation_id = num_res["activation_id"]
        raw_phone = num_res["phone"]
        
        # Format phone (keep last 10 digits for Indian numbers)
        phone = raw_phone
        if phone.startswith("91") and len(phone) == 12:
            phone = phone[2:]
            
        print(f"📞 Got Number: {phone} (ID: {activation_id})")
        print(f"📞 नया नंबर: {phone}\n🌐 API Mode: Background requests initialization...")

        success_count = 0
        sms_received_count = 0
        activation_cancelled = False
        used_otps = set()

        try:
            for host_url in urls:
                if not is_automation_running.get(user_id, False):
                    break
                    
                if not host_url.startswith("http"):
                    formatted_url = "https://" + host_url
                else:
                    formatted_url = host_url

                try:
                    print(f"🌐 (API Mode) Registering on {host_url}...")
                    
                    # 1. Start Session & Fetch CSRF
                    session = requests.Session()
                    session.headers.update({
                        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                        "X-Requested-With": "XMLHttpRequest",
                        "Origin": formatted_url,
                        "Referer": formatted_url + "/"
                    })
                    
                    r_home = session.get(formatted_url, timeout=15)
                    
                    token = None
                    match = re.search(r'name="csrf-token"\s+content="([^"]+)"', r_home.text)
                    if not match:
                        match = re.search(r'value="([^"]+)"[^>]*name="_token"', r_home.text)
                    if match:
                        token = match.group(1)
                        
                    if not token:
                        raise Exception("CSRF token not found on site homepage")
                        
                    print(f"🔑 CSRF Token extracted: {token[:15]}...")

                    # 2. GENERATE UNIQUE USERNAME
                    username = generate_username()
                    email = f"{username}@gmail.com"

                    # 3. POST registration form to trigger OTP send
                    payload = {
                        "username": username,
                        "email": email,
                        "password": "XnX@1",
                        "phone": phone,
                        "otp": "",
                        "_token": token
                    }
                    headers = {
                        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"
                    }
                    
                    reg_endpoint = f"{formatted_url.rstrip('/')}/register"
                    print(f"📡 Sending registration trigger payload for username '{username}'...")
                    r_send = session.post(reg_endpoint, data=payload, headers=headers, timeout=15)
                    
                    try:
                        res_send = r_send.json()
                        print(f"📩 Server response: {res_send}")
                    except Exception as parse_err:
                        raise Exception(f"Failed to parse registration response JSON: {r_send.text[:150]}")
                        
                    # Check if trigger succeeded
                    is_sent = ("success" in str(res_send.get("message_class", "")).lower() or 
                               "sent" in str(res_send.get("message", "")).lower() or
                               res_send.get("status") == 205)
                               
                    if not is_sent:
                        raise Exception(f"Registration trigger rejected: {res_send.get('message', 'Unknown error')}")

                    print(f"📡 Form submitted successfully. Waiting for OTP...")

                    # 4. POLL OTP DOCTOR FOR OTP
                    start_time = time.time()
                    extracted_otp = None
                    timeout = 60
                    
                    while time.time() - start_time < timeout:
                        if not is_automation_running.get(user_id, False):
                            break
                            
                        status_res = get_otp_status(activation_id, api_key)
                        print(f"👀 Polling OTP Status: {status_res}")
                        
                        if status_res.startswith("STATUS_OK:"):
                            sms_text = status_res.split(":", 1)[1]
                            extracted_otp = extract_otp_code(sms_text, used_otps)
                            if extracted_otp:
                                print(f"🎯 New OTP Retrieved for {host_url}: {extracted_otp}")
                                used_otps.add(extracted_otp)
                                break
                        elif status_res == "STATUS_CANCEL":
                            raise Exception("Activation cancelled from provider side.")
                        elif status_res == "NO_ACTIVATION":
                            raise Exception("Activation session not found on provider.")
                        
                        time.sleep(3)

                    if not extracted_otp:
                        raise Exception("OTP timeout / OTP not received")

                    # We received an SMS! Increment the received count
                    sms_received_count += 1

                    # 5. SUBMIT VERIFICATION WITH OTP
                    payload["otp"] = extracted_otp
                    print(f"🔢 Verifying OTP {extracted_otp} on {host_url}...")
                    r_verify = session.post(reg_endpoint, data=payload, headers=headers, timeout=15)
                    
                    try:
                        res_verify = r_verify.json()
                        print(f"📩 Verification response: {res_verify}")
                    except Exception as parse_err:
                        raise Exception(f"Failed to parse verification response JSON: {r_verify.text[:150]}")
                        
                    is_verified = (res_verify.get("status") == 200 or 
                                   "success" in str(res_verify.get("message_class", "")).lower() or 
                                   res_verify.get("statusCode") == 200 or
                                   "already" in str(res_verify.get("message", "")).lower())
                                   
                    if not is_verified:
                        raise Exception(f"Verification rejected: {res_verify.get('message', 'Invalid OTP')}")

                    # 6. API REQUEST & SERVER RESPONSE VALIDATION LOOP (Final confirmation)
                    from urllib.parse import urlparse
                    parsed = urlparse(formatted_url)
                    target_api_url = f"{parsed.scheme}://{parsed.netloc}/send_otp_touser"
                    success_received = False
                    retry_count = 0
                    max_api_retries = 5

                    while not success_received and is_automation_running.get(user_id, False) and retry_count < max_api_retries:
                        retry_count += 1
                        random_api_phone = str(random.randint(100000000000000, 999999999999999))
                        print(f"📡 Sending API Request (Attempt {retry_count}/{max_api_retries}) for Phone: {random_api_phone}...")

                        api_payload = {
                            "_token": token,
                            "phone": random_api_phone,
                            "url": host_url
                        }
                        
                        r_api = session.post(target_api_url, data=api_payload, headers=headers, timeout=15)
                        try:
                            server_data = r_api.json()
                            print(f"📩 Server Data Received: {server_data}")
                        except:
                            server_data = {}

                        # Normalize list of dicts response format
                        if isinstance(server_data, list) and len(server_data) > 0:
                            actual_res = server_data[0]
                        elif isinstance(server_data, dict):
                            actual_res = server_data
                        else:
                            actual_res = {}

                        if actual_res.get("statusCode") == 200 or "success" in str(actual_res.get("message_class", "")).lower():
                            print("🎯 Success Response Matched!")
                            success_received = True
                            break
                        else:
                            print("⚠️ Server response failed or mismatched. Retrying in 3 seconds...")
                            time.sleep(3)

                    if success_received:
                        save_user(host_url, username, "XnX@1", phone)
                        success_count += 1
                        print(f"✅ {host_url} पर पंजीकरण सफल!")
                    else:
                        raise Exception("API verification failed (did not get success response from server)")

                except Exception as domain_err:
                    print(f"⚠️ Registration failed for {host_url}: {domain_err}")
                
                # At the end of the domain loop iteration, prepare for next SMS if we received one
                if sms_received_count > 0 and host_url != urls[-1]:
                    try:
                        print("📡 Requesting next SMS status=3 to keep channel open...")
                        set_otp_status(activation_id, 3, api_key)
                        time.sleep(3)
                    except Exception as status_err:
                        print(f"⚠️ Failed to request next SMS (status=3): {status_err}")

            if success_count == 0:
                print("🛑 No domains succeeded. Cancelling number to get refund...")
                set_otp_status(activation_id, 8, api_key)
                activation_cancelled = True
                
                # Rotate candidate list so we try a different service ID next time
                if selected_service_id == "auto" and used_service_id in candidate_services:
                    candidate_services.remove(used_service_id)
                    candidate_services.append(used_service_id)
                    print(f"🔄 Rotated failed service {used_service_id} to end. New priority: {candidate_services}")
                    
                print("⚠️ सभी डोमेन पर पंजीकरण विफल। नंबर रद्द कर दिया गया है। 5 सेकंड में पुनः प्रयास शुरू होगा...")
            else:
                print(f"🎉 Successfully registered on {success_count} domains using number {phone}!")
                print(f"🎉 चक्र पूरा! {success_count} डोमेन पर पंजीकरण सफल। अगला लूप 5 सेकंड में चालू हो रहा है...")
            time.sleep(5)

        except Exception as e:
            print("❌ ERROR IN FLOW:", e)

            if success_count == 0 and not activation_cancelled:
                try:
                    set_otp_status(activation_id, 8, api_key)
                except:
                    pass
                    
                # Rotate candidate list so we try a different service ID next time
                if selected_service_id == "auto" and used_service_id in candidate_services:
                    candidate_services.remove(used_service_id)
                    candidate_services.append(used_service_id)
                    print(f"🔄 Rotated failed service {used_service_id} to end. New priority: {candidate_services}")

            time.sleep(5)


# ---------------- HTTP SERVER INIT ----------------


def start_http_server():
    try:
        port = int(os.getenv("PORT", 8080))
        server = HTTPServer(("0.0.0.0", port), DashboardHandler)
        add_log(f"🌐 Dashboard Web Server running on port {port}")
        server.serve_forever()
    except Exception as e:
        add_log(f"❌ Error starting HTTP server: {e}")


# ---------------- MAIN ----------------
def main():
    # Start Web Dashboard
    start_http_server()


if __name__ == "__main__":
    main()